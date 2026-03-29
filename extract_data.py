#!/usr/bin/env python3
"""Extract data from the Excel workbook and produce data.json for the HTML demo."""

import json
import math
import openpyxl
import numpy as np
from datetime import datetime

EXCEL_PATH = "Devoir 2 - MQ2 - lymoi.xlsx"
OUTPUT_PATH = "data.json"

# Canonical ticker order (matches TICKER sheet)
TICKERS = ["AAPL", "C", "F", "HPQ", "JNJ", "LLY", "LOW", "MO", "MRK", "WMT"]

# Sheet name mapping (Excel sheet names differ from tickers)
SHEET_MAP = {
    "AAPL": "AAPPL",
    "C": "C",
    "F": "F",
    "HPQ": "HPQ",
    "JNJ": "J&J",
    "LLY": "LLY",
    "LOW": "LOW",
    "MO": "MO",
    "MRK": "MRK",
    "WMT": "WMT",
}

STOCK_NAMES = {
    "AAPL": "Apple Inc",
    "C": "Citigroup Inc",
    "F": "Ford Motor Co",
    "HPQ": "Hewlett-Packard Co",
    "JNJ": "Johnson & Johnson",
    "LLY": "Eli Lilly & Co",
    "LOW": "Lowe's Cos Inc",
    "MO": "Altria Group Inc",
    "MRK": "Merck & Co Inc",
    "WMT": "Wal-Mart Stores Inc",
}

EXCHANGES = {
    "AAPL": "NASDAQ",
    "C": "NYSE",
    "F": "NYSE",
    "HPQ": "NYSE",
    "JNJ": "NYSE",
    "LLY": "NYSE",
    "LOW": "NYSE",
    "MO": "NYSE",
    "MRK": "NYSE",
    "WMT": "NYSE",
}


def extract_stock_sheet(wb, ticker):
    """Extract weekly dates and adjusted prices from a stock sheet."""
    sheet_name = SHEET_MAP[ticker]
    ws = wb[sheet_name]
    dates = []
    adj_prices = []
    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row, 1).value  # col A = date
        adj_price = ws.cell(row, 3).value  # col C = adjusted last price
        if date_val is None or adj_price is None:
            continue
        if isinstance(date_val, datetime):
            dates.append(date_val.strftime("%Y-%m-%d"))
        else:
            continue
        adj_prices.append(float(adj_price))
    return dates, adj_prices


def extract_volatility(wb, ticker):
    """Extract annualized volatility from cell G2 of the stock sheet."""
    sheet_name = SHEET_MAP[ticker]
    ws = wb[sheet_name]
    vol = ws["G2"].value
    return float(vol) if vol is not None else None


def extract_initial_prices(wb):
    """Extract initial share prices from the TICKER sheet."""
    ws = wb["TICKER"]
    prices = {}
    for row in range(2, 12):
        ticker = ws.cell(row, 2).value  # col B = ticker
        price = ws.cell(row, 3).value  # col C = initial price
        if ticker and price:
            prices[ticker] = float(price)
    return prices


def compute_log_returns(prices):
    """Compute log returns from a price series."""
    returns = []
    for i in range(1, len(prices)):
        if prices[i] > 0 and prices[i - 1] > 0:
            returns.append(math.log(prices[i] / prices[i - 1]))
        else:
            returns.append(None)
    return returns


def compute_correlation_matrix(all_returns):
    """Compute 10x10 correlation matrix from aligned return series.

    all_returns: dict of ticker -> list of returns (same length, common dates)
    Returns: 10x10 numpy array, verified symmetric.
    """
    n_stocks = len(TICKERS)
    # Build matrix: rows = observations, cols = stocks
    n_obs = len(all_returns[TICKERS[0]])
    matrix = np.zeros((n_obs, n_stocks))
    for j, ticker in enumerate(TICKERS):
        for i in range(n_obs):
            matrix[i, j] = all_returns[ticker][i]

    corr = np.corrcoef(matrix, rowvar=False)

    # Verify symmetry
    assert np.allclose(corr, corr.T), "Correlation matrix is not symmetric!"
    # Verify diagonal
    assert np.allclose(np.diag(corr), 1.0), "Diagonal entries are not 1.0!"
    # Verify range
    assert np.all(corr >= -1.0) and np.all(corr <= 1.0), "Values outside [-1, 1]!"

    return corr


def main():
    print(f"Loading {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # --- Extract per-stock data ---
    all_dates = {}
    all_prices = {}
    volatilities = {}
    initial_prices = extract_initial_prices(wb)

    for ticker in TICKERS:
        print(f"  Extracting {ticker} ({SHEET_MAP[ticker]})...")
        dates, prices = extract_stock_sheet(wb, ticker)
        all_dates[ticker] = dates
        all_prices[ticker] = prices
        volatilities[ticker] = extract_volatility(wb, ticker)
        print(f"    {len(dates)} rows, vol={volatilities[ticker]:.4f}")

    # --- Find common date intersection ---
    date_sets = [set(all_dates[t]) for t in TICKERS]
    common_dates = sorted(set.intersection(*date_sets))
    print(f"\n  Common dates: {len(common_dates)} (from {common_dates[0]} to {common_dates[-1]})")

    # --- Build aligned price series on common dates ---
    aligned_prices = {}
    for ticker in TICKERS:
        date_to_price = dict(zip(all_dates[ticker], all_prices[ticker]))
        aligned_prices[ticker] = [date_to_price[d] for d in common_dates]

    # --- Compute log returns on common dates ---
    aligned_returns = {}
    for ticker in TICKERS:
        returns = compute_log_returns(aligned_prices[ticker])
        # Filter out None values (shouldn't happen on aligned data, but safety)
        aligned_returns[ticker] = returns
    # Return dates correspond to dates[1:] (first date has no return)
    return_dates = common_dates[1:]

    # --- Compute 10x10 correlation matrix ---
    print("\n  Computing correlation matrix (log returns, common dates)...")
    corr_matrix = compute_correlation_matrix(aligned_returns)
    print("  Correlation matrix: symmetric ✓, diagonal=1 ✓, range [-1,1] ✓")

    # --- Verify Cholesky decomposition succeeds (positive definite) ---
    try:
        L = np.linalg.cholesky(corr_matrix)
        print("  Cholesky decomposition: success ✓ (positive definite)")
    except np.linalg.LinAlgError:
        print("  WARNING: Cholesky failed! Matrix is not positive definite.")
        L = None

    # --- S0 prices (last available adjusted price, Oct 28 2016) ---
    s0_prices = {}
    for ticker in TICKERS:
        s0_prices[ticker] = aligned_prices[ticker][-1]
    last_date = common_dates[-1]
    print(f"\n  S0 prices (as of {last_date}):")
    for t in TICKERS:
        print(f"    {t}: {s0_prices[t]:.2f}")

    # --- Adjusted initial prices at trade date (Sept 26, 2011) ---
    # The TICKER sheet has RAW market prices, but price_history uses adjusted prices.
    # We need the adjusted price at the trade date for consistent payoff calculation.
    trade_date = "2011-09-26"
    trade_idx = next(i for i, d in enumerate(common_dates) if d >= trade_date)
    trade_date_actual = common_dates[trade_idx]
    adj_initial_prices = {}
    print(f"\n  Adjusted initial prices (trade date {trade_date_actual}):")
    for t in TICKERS:
        adj_initial_prices[t] = aligned_prices[t][trade_idx]
        print(f"    {t}: adj={adj_initial_prices[t]:.4f}  (raw from TICKER: {initial_prices.get(t)})")

    # --- Weekly std devs and annualized vols ---
    weekly_stds = {}
    for ticker in TICKERS:
        weekly_stds[ticker] = float(np.std(aligned_returns[ticker], ddof=1))

    # --- Build output JSON ---
    output = {
        "meta": {
            "source": EXCEL_PATH,
            "extracted_at": datetime.now().isoformat(),
            "common_date_range": [common_dates[0], common_dates[-1]],
            "n_common_dates": len(common_dates),
            "n_returns": len(return_dates),
            "return_type": "log_returns",
        },
        "product": {
            "name": "Annual Digital Yield Generator Certificates of Deposit",
            "cusip": "05573J AX2",
            "term_years": 6,
            "trade_date": "2011-09-26",
            "settlement_date": "2011-09-29",
            "issue_date": "2011-09-29",
            "maturity_date": "2017-09-29",
            "denomination": 1000,
            "digital_coupon": 0.065,
            "floor": -0.30,
            "coupon_dates": [
                "2012-09-28",
                "2013-09-30",
                "2014-09-30",
                "2015-09-30",
                "2016-09-30",
                "2017-09-29",
            ],
            "default_eval_date": "2016-10-31",
            "default_end_date": "2017-09-29",
            "default_risk_free_rate": 0.0125,
            "delta_t": 11 / 12,
        },
        "stocks": [
            {
                "ticker": t,
                "name": STOCK_NAMES[t],
                "exchange": EXCHANGES[t],
                "initial_price_raw": initial_prices.get(t),
                "initial_price_adj": round(adj_initial_prices[t], 4),
                "initial_price_adj_date": trade_date_actual,
                "weight": 0.10,
                "s0_price": round(s0_prices[t], 4),
                "s0_date": last_date,
                "annualized_vol": round(volatilities[t], 6),
                "weekly_std": round(weekly_stds[t], 6),
            }
            for t in TICKERS
        ],
        "correlation_matrix": {
            "tickers": TICKERS,
            "matrix": [[round(corr_matrix[i][j], 6) for j in range(10)] for i in range(10)],
        },
        "cholesky_L": {
            "tickers": TICKERS,
            "matrix": [[round(L[i][j], 6) for j in range(10)] for i in range(10)]
            if L is not None
            else None,
        },
        "price_history": {
            "dates": common_dates,
            "prices": {t: [round(p, 4) for p in aligned_prices[t]] for t in TICKERS},
        },
    }

    # --- Write JSON ---
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2)
    print(f"\n✓ Wrote {OUTPUT_PATH} ({len(json.dumps(output)) / 1024:.0f} KB)")


if __name__ == "__main__":
    main()
